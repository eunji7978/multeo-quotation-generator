from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from datetime import datetime
import os

def create_quotation(recipient_name, items, output_path="quotation.xlsx"):
    """
    Generates a quotation Excel file.
    
    Args:
        recipient_name (str): Name of the recipient (e.g. "OOO 귀하")
        items (list): List of dicts with keys: 'name', 'quantity', 'unit_price'
        output_path (str): Path to save the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"

    # --- Styles ---
    border_thin = Side(style='thin')
    border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    font_title = Font(size=20, bold=True, underline='single')
    font_header = Font(bold=True)
    
    # --- Layout Setup ---
    # Columns: A(No), B(Product), C(Spec), D(Qty), E(Unit Price), F(Supply Price), G(Total), H(Tax), I(Note)
    ws.column_dimensions['A'].width = 6   # No
    ws.column_dimensions['B'].width = 35  # Item Name (Increased)
    ws.column_dimensions['C'].width = 10  # Spec
    ws.column_dimensions['D'].width = 8   # Qty
    ws.column_dimensions['E'].width = 15  # Unit Price
    ws.column_dimensions['F'].width = 15  # Supply Price
    ws.column_dimensions['G'].width = 15  # Sum
    ws.column_dimensions['H'].width = 10  # Tax Note
    ws.column_dimensions['I'].width = 20  # Remarks / Jongmok (Increased)

    def set_border(ws, cell_range):
        """Apply border to a range of cells"""
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.border = border_all

    # --- Header Section ---
    ws.merge_cells('B2:I3') # Title centered over main area
    cell = ws['B2']
    cell.value = "견  적  서"
    cell.font = font_title
    cell.alignment = align_center
    cell.border = Border(bottom=Side(style='double'))

    # Date
    today = datetime.now()
    date_str = today.strftime("%Y년 %m월 %d일")
    # Shift to B. B is wide (35), so maybe no merge needed, or merge B:D safely.
    ws.merge_cells('B5:D5')
    ws['B5'] = date_str
    ws['B5'].alignment = align_left
    
    # Recipient
    ws.merge_cells('B7:D7')
    ws['B7'] = f"{recipient_name}   귀하"
    ws['B7'].font = Font(size=14, bold=True)
    ws['B7'].alignment = align_left
    ws['B7'].border = Border(bottom=border_thin) # Underline effect
    
    ws.merge_cells('B9:D9')
    ws['B9'] = "아래와 같이 견적합니다."
    ws['B9'].alignment = align_left

    # Supplier Info Box (Fixed)
    # Range E5:I9
    
    # "Supplier" Vertical Label
    ws.merge_cells('E5:E9')
    ws['E5'] = "공급자"
    ws['E5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Registration No
    ws.merge_cells('F5:G5')
    ws['F5'] = "등 록 번 호"
    ws['F5'].alignment = align_center
    
    ws.merge_cells('H5:I5')
    ws['H5'] = "692-27-01163"
    ws['H5'].alignment = align_center
    
    # Company Name & Name
    # Row 6: F=Label, G=Val, H=Label, I=Val
    ws['F6'] = "상호(법인명)"
    ws['F6'].alignment = align_center
    
    ws['G6'] = "물터 (multeo)"
    ws['G6'].alignment = align_center
    
    ws['H6'] = "성  명"
    ws['H6'].alignment = align_center
    
    ws['I6'] = "김예지  (인)" 
    ws['I6'].alignment = align_center
    
    # Address
    ws['F7'] = "사업장주소"
    ws['F7'].alignment = align_center
    
    ws.merge_cells('G7:I7')
    ws['G7'] = "서울시 성동구 행당동 138-7 202호"
    ws['G7'].alignment = align_left
    
    # Business Type
    ws['F8'] = "업    태"
    ws['F8'].alignment = align_center
    
    ws['G8'] = "소매업"
    ws['G8'].alignment = align_center
    
    ws['H8'] = "종  목"
    ws['H8'].alignment = align_center
    
    ws['I8'] = "공예, 시각디자인"
    ws['I8'].alignment = align_center
    
    # Phone
    ws['F9'] = "전 화 번 호"
    ws['F9'].alignment = align_center
    
    ws.merge_cells('G9:I9')
    ws['G9'] = "010-4341-7978"
    ws['G9'].alignment = align_left

    # Apply borders to the whole Supplier Box E5:I9
    set_border(ws, 'E5:I9')


    # --- Total Amount Section ---
    total_supply_price = 0
    for item in items:
        unit_price = item['unit_price']
        supply_price = int(unit_price * 0.6)
        qty = item['quantity']
        total_supply_price += (supply_price * qty)

    grand_total = total_supply_price
    
    # Row 11: Total Amount
    # Merge B11:I12 (Align with Item Table which starts at B)
    ws.merge_cells('B11:I12') 
    
    total_text = f"합 계 금 액   (공급가액VAT포함)          KRW {grand_total:,}"
    ws['B11'] = total_text
    ws['B11'].font = Font(size=16, bold=True)
    ws['B11'].alignment = align_center
    
    # Apply border to B11:I12
    set_border(ws, 'B11:I12')

    # --- Items Header ---
    headers = ["품  명", "규  격", "수 량", "단  가", "공 급 가 액", "합  계", "세 액", "비  고"]
    header_row = 14
    ws.merge_cells('B14:B14') 
    
    ws['B14'] = "품  명"
    ws['C14'] = "규  격"
    ws['D14'] = "수 량"
    ws['E14'] = "단  가"
    ws['F14'] = "공 급 가 액"
    ws['G14'] = "합  계"
    ws['H14'] = "세 액"
    ws['I14'] = "비  고"
    
    # Style logic for header: Gray background, Center, Border
    for col_idx in range(2, 10): # B to I
        cell = ws.cell(row=header_row, column=col_idx)
        cell.alignment = align_center
        cell.font = font_header
        cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    # Apply border to header row B14:I14
    set_border(ws, 'B14:I14')

    # --- Items Data ---
    start_row = 15
    current_row = start_row
    
    for item in items:
        # Set row height for data rows
        ws.row_dimensions[current_row].height = 25
        
        name = item['name']
        qty = item['quantity']
        unit_price = item['unit_price']
        supply_price = int(unit_price * 0.6)
        row_sum = supply_price * qty
        
        c_name = ws.cell(row=current_row, column=2, value=name)
        c_name.alignment = align_center # User requested Center
        
        ws.cell(row=current_row, column=3, value="") # Spec
        
        c_qty = ws.cell(row=current_row, column=4, value=qty)
        c_qty.alignment = align_center
        
        c_unit = ws.cell(row=current_row, column=5, value=unit_price)
        c_unit.number_format = '#,##0'
        c_unit.alignment = align_center # User requested Center
        
        c_supply = ws.cell(row=current_row, column=6, value=supply_price)
        c_supply.number_format = '#,##0'
        c_supply.alignment = align_center # User requested Center
        
        c_sum = ws.cell(row=current_row, column=7, value=row_sum)
        c_sum.number_format = '#,##0'
        c_sum.alignment = align_center # User requested Center
        
        c_tax = ws.cell(row=current_row, column=8, value="포함")
        c_tax.alignment = align_center
        
        ws.cell(row=current_row, column=9, value="") # Note

        current_row += 1
        
    # Fill remaining rows up to Header Row for Footer (footer_row)
    # The footer is at footer_row (31). We want to fill up to 30.
    footer_row = 31
    for r in range(current_row, footer_row):
        ws.row_dimensions[r].height = 25
        for c in range(2, 10):
            ws.cell(row=r, column=c, value="")

    # Apply borders to the entire Item Table Section
    item_table_range = f"B{start_row}:I{footer_row - 1}"
    set_border(ws, item_table_range)

    # --- Footer ---
    ws.merge_cells(f'B{footer_row}:I{footer_row}')
    ws[f'B{footer_row}'] = "신한 110-540-155149 김예지(물터)"
    ws[f'B{footer_row}'].alignment = Alignment(horizontal='center', vertical='bottom')
    ws[f'B{footer_row}'].font = Font(bold=True)
    
    f_total_row = footer_row + 1
    ws[f'B{f_total_row}'] = "합계금액"
    ws[f'B{f_total_row}'].alignment = align_center
    
    ws.merge_cells(f'C{f_total_row}:F{f_total_row}')
    
    ws[f'G{f_total_row}'] = grand_total
    ws[f'G{f_total_row}'].number_format = '#,##0'
    
    ws[f'H{f_total_row}'] = "포함"
    ws[f'H{f_total_row}'].alignment = align_center
    
    # Apply borders to Footer Table
    set_border(ws, f'B{footer_row}:I{footer_row}')
    set_border(ws, f'B{f_total_row}:I{f_total_row}')

    # --- Stamp Insertion ---
    stamp_path = "assets/stamp.png"
    if os.path.exists(stamp_path):
        try:
            img = Image(stamp_path)
            # size = 1.08" x 1.39" (104x133px)
            img.width = 104
            img.height = 133
            
            # Position: Overlap "(인)"
            # User liked Offset 80 for Width 53 (Center ~106px).
            # For Width 104, to keep Center ~106px, Offset must be ~55px.
            # Vertical: User requested "Slightly Up" again (5 -> 0). Top align.
            col_idx = 8 
            row_idx = 3
            
            size = XDRPositiveSize2D(pixels_to_EMU(104), pixels_to_EMU(133))
            marker = AnchorMarker(col=col_idx, colOff=pixels_to_EMU(55), row=row_idx, rowOff=pixels_to_EMU(0))
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            
            ws.add_image(img)
        except Exception as e:
            # Fallback
            print(f"Stamp error (complex anchor): {e}")
            try:
                img = Image(stamp_path)
                img.width = 45
                img.height = 45
                ws.add_image(img, 'I6')
            except Exception as e2:
                print(f"Could not load stamp: {e2}")

    wb.save(output_path)
    print(f"Quotation saved to {output_path}")

if __name__ == "__main__":
    # Test
    test_items = [
        {"name": "실버볼 슬림컵", "quantity": 9, "unit_price": 70000},
        {"name": "실버볼 컵", "quantity": 6, "unit_price": 60000}
    ]
    create_quotation("레퍼토리 성수", test_items, "test_quotation.xlsx")
