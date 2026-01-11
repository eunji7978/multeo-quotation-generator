from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from datetime import datetime
import os

def create_calibration_file(offset, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"

    # --- Styles & Layout (Simplified for speed, but matching needed structure) ---
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['I'].width = 20

    # Header
    ws.merge_cells('B2:I3')
    ws['B2'] = f"도장 위치 테스트 (Offset: {offset})"
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].font = Font(size=14, bold=True)

    # Supplier Box structure
    ws.merge_cells('E5:E9')
    # ws.merge_cells('F6:I6') # Removed incorrect merge
    # Actually, let's replicate the specific cell I6 structure
    # I6 is Column 9 (Index 8 in 0-based).
    # "김예지 (인)"
    
    ws['I6'] = "김예지 (인)"
    ws['I6'].alignment = Alignment(horizontal='center', vertical='center')
    ws['I6'].border = Border(bottom=Side(style='thin'))

    # Stamp Insertion
    stamp_path = "assets/stamp.png"
    if os.path.exists(stamp_path):
        img = Image(stamp_path)
        img.width = 53
        img.height = 75
        
        col_idx = 8 
        row_idx = 3 # Fixed Row (compensated)
        
        size = XDRPositiveSize2D(pixels_to_EMU(53), pixels_to_EMU(75))
        marker = AnchorMarker(col=col_idx, colOff=pixels_to_EMU(offset), row=row_idx, rowOff=pixels_to_EMU(17))
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        
        ws.add_image(img)

    wb.save(filename)
    print(f"Generated {filename}")

if __name__ == "__main__":
    # Generate 3 variations
    offsets = [75, 80, 85]
    for off in offsets:
        create_calibration_file(off, f"Test_Stamp_{off}.xlsx")
